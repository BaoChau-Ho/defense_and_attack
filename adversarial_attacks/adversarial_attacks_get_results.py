import torch
import argparse
import sys
from os import system
import time
import copy
import numpy as np
from torch import optim
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader, TensorDataset
from torchvision import datasets
from torchvision.models.inception import InceptionAux
from torchvision.models import mobilenet_v2, resnet50, alexnet, resnet101, densenet121, inception_v3, DenseNet121_Weights, ResNet101_Weights, ResNet50_Weights, MobileNet_V2_Weights, AlexNet_Weights, Inception_V3_Weights
from torchvision import transforms
from pytorchtools import EarlyStopping
from utils import Logger
from pynvml import *
import torchattacks

import json
import os
import openpyxl as yxl
import numpy as np
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

def make_directory(dirs):
    for dir in dirs:
        if(not os.path.exists(dir)):
            os.makedirs(dir)

parser = argparse.ArgumentParser()
parser.add_argument("--index_model",type=int,default=0)
parser.add_argument("--index_dataset",type=int,default=0)
parser.add_argument("--father_directory", type=str,default=os.getcwd())
parser.add_argument("--number_of_imgs",type=int,default=10)
args=parser.parse_args()

device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
device1 = torch.device("cuda:0")

bs = 32

models = ['resnet50_CIFA100', 'resnet101_CIFA100', 'mobilenet_v2_CIFA100', 'alexnet_CIFA100', 'densenet121_CIFA100','inception_v3_CIFA100',
          'resnet50_CIFA10', 'resnet101_CIFA10', 'mobilenet_v2_CIFA10', 'alexnet_CIFA10', 'densenet121_CIFA10', 'inception_v3_CIFA10',
          'resnet50_MNIST', 'resnet101_MNIST', 'mobilenet_v2_MNIST', 'alexnet_MNIST', 'densenet121_MNIST', 'inception_v3_MNIST',
          'resnet50_IMGNET', 'resnet101_IMGNET','mobilenet_v2_IMGNET','alexnet_IMGNET','densenet121_IMGNET','inception_v3_IMGNET']
adver_methods = ['FGSM','BIM','PGD','CW_L2']
all_datasets = ['cifar100','cifar10','mnist','imgnet']

father_directory = args.father_directory

print(father_directory)

# 0-5: cifar100
# 6-11: cifar10
# 12-17: mnist
# 18-23: imgnet
model_name = models[args.index_model] 
dataset_name = all_datasets[args.index_dataset]
model_path = os.path.join(father_directory, "models", "{}".format(model_name), "weights_{}_best.h5".format(model_name))
data_path = os.path.join(father_directory, "datasets_{}".format(dataset_name))
dst_dir = os.path.join(father_directory, "adversarial_dataset_{}".format(model_name))
label_path = os.path.join(father_directory, "{}-labels.json".format(dataset_name))
excel_dir = os.path.join(father_directory, "excel_result", "{}".format(dataset_name),"{}".format(model_name))
img_dst_dir = os.path.join(father_directory, "Image_and_Noise", "{}".format(dataset_name), "{}".format(model_name))
print(model_name)
print(dataset_name)

make_directory([dst_dir])

mean = (0.4914, 0.4822, 0.4465)
std = (0.247, 0.243, 0.261)
    
def transforms_data_cifa100(width=224, height=224):
    transforms_dict = {
        'train':
        transforms.Compose([
            transforms.Resize((width, height)),
            #transforms.RandomAffine(0,scale=(0.8, 1.2)),
            transforms.RandomAffine(0, shear=10, scale=(0.8, 1.2)),
            transforms.RandomHorizontalFlip(), # ngang
            transforms.ToTensor(),
            transforms.Normalize(mean , std)
        ]),
        'validation':
        transforms.Compose([
            transforms.Resize((width, height)),
            transforms.ToTensor(),
            transforms.Normalize(mean, std)
        ]),
    }
    return transforms_dict
    
def transforms_data_mnist(width=224, height=224):
    transforms_dict = {
        'train':
        transforms.Compose([
            transforms.Grayscale(3),
            transforms.Resize((width, height)),
            #transforms.RandomAffine(0,scale=(0.8, 1.2)),
            transforms.RandomAffine(0, shear=10, scale=(0.8, 1.2)),
            transforms.RandomHorizontalFlip(), # ngang
            transforms.ToTensor(),
            transforms.Normalize(mean , std)
        ]),
        'validation':
        transforms.Compose([
            transforms.Grayscale(3),
            transforms.Resize((width, height)),
            transforms.ToTensor(),
            transforms.Normalize(mean, std)
        ]),
    }
    return transforms_dict

def transforms_data_cifa10(width=224, height=224):
    transforms_dict = {
        'train':
        transforms.Compose([
            transforms.Resize((width, height)),
            #transforms.RandomAffine(0,scale=(0.8, 1.2)),
            transforms.RandomAffine(0, shear=10, scale=(0.8, 1.2)),
            transforms.RandomHorizontalFlip(), # ngang
            transforms.ToTensor(),
            transforms.Normalize(mean , std)
        ]),
        'validation':
        transforms.Compose([
            transforms.Resize((width, height)),
            transforms.ToTensor(),
            transforms.Normalize(mean, std)
        ]),
    }
    return transforms_dict

def transforms_data_imgnet(width=224, height=224):
    transforms_dict = {
        'train':
        transforms.Compose([
            transforms.Resize((width, height)),
            #transforms.RandomAffine(0,scale=(0.8, 1.2)),
            transforms.RandomAffine(0, shear=10, scale=(0.8, 1.2)),
            transforms.RandomHorizontalFlip(), # ngang
            transforms.ToTensor(),
            transforms.Normalize(mean , std)
        ]),
        'validation':
        transforms.Compose([
            transforms.Resize((width, height)),
            transforms.ToTensor(),
            transforms.Normalize(mean, std)
        ]),
    }
    return transforms_dict

def initModel(path = None):
    if(model_name == 'mobilenet_v2_CIFA100'):
        model = mobilenet_v2(weights = MobileNet_V2_Weights.DEFAULT)
        in_features = model.classifier[1].in_features
        model.classifier[1] = torch.nn.Linear(in_features=in_features, out_features=100)
    if(model_name == 'resnet50_CIFA100'):
        model = resnet50(weights = ResNet50_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features=in_features, out_features = 100)
    if(model_name == 'alexnet_CIFA100'):
        model = alexnet(weights = AlexNet_Weights.DEFAULT)
        in_features = model.classifier[6].in_features
        model.classifier[6] = torch.nn.Linear(in_features = in_features, out_features = 100)
    if(model_name == 'resnet101_CIFA100'):
        model = resnet101(weights = ResNet101_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 100)
    if(model_name == 'densenet121_CIFA100'):
        model = densenet121(weights = DenseNet121_Weights.DEFAULT)
        in_features = model.classifier.in_features
        model.classifier = torch.nn.Linear(in_features=in_features, out_features=100)
    if(model_name == 'inception_v3_CIFA100'):
        model = inception_v3(weights = Inception_V3_Weights.DEFAULT)
        in_features = model.fc.in_features
        in_features2 = 768
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 100)
        model.AuxLogits = InceptionAux(in_channels = in_features2, num_classes = 100)
    if(model_name == 'resnet50_CIFA10'):
        model = resnet50(weights = ResNet50_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features=in_features, out_features = 10)
    if(model_name == 'resnet101_CIFA10'):
        model = resnet101(weights = ResNet101_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 10)
    if(model_name == 'mobilenet_v2_CIFA10'):
        model = mobilenet_v2(weights = MobileNet_V2_Weights.DEFAULT)
        in_features = model.classifier[1].in_features
        model.classifier[1] = torch.nn.Linear(in_features=in_features, out_features=10)       
    if(model_name == 'alexnet_CIFA10'):
        model = alexnet(weights = AlexNet_Weights.DEFAULT)
        in_features = model.classifier[6].in_features
        model.classifier[6] = torch.nn.Linear(in_features = in_features, out_features = 10) 
    if(model_name == 'densenet121_CIFA10'):
        model = densenet121(weights = DenseNet121_Weights.DEFAULT)
        in_features = model.classifier.in_features
        model.classifier = torch.nn.Linear(in_features=in_features, out_features=10)     
    if(model_name == 'inception_v3_CIFA10'):
        model = inception_v3(weights = Inception_V3_Weights.DEFAULT)
        in_features = model.fc.in_features
        in_features2 = 768
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 10)
        model.AuxLogits = InceptionAux(in_channels = in_features2, num_classes = 10)  
    if(model_name == 'resnet50_MNIST'):
        model = resnet50(weights = ResNet50_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features=in_features, out_features = 10)
    if(model_name == 'resnet101_MNIST'):
        model = resnet101(weights = ResNet101_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 10)
    if(model_name == 'mobilenet_v2_MNIST'):
        model = mobilenet_v2(weights = MobileNet_V2_Weights.DEFAULT)
        in_features = model.classifier[1].in_features
        model.classifier[1] = torch.nn.Linear(in_features=in_features, out_features=10)       
    if(model_name == 'alexnet_MNIST'):
        model = alexnet(weights = AlexNet_Weights.DEFAULT)
        in_features = model.classifier[6].in_features
        model.classifier[6] = torch.nn.Linear(in_features = in_features, out_features = 10) 
    if(model_name == 'densenet121_MNIST'):
        model = densenet121(weights = DenseNet121_Weights.DEFAULT)
        in_features = model.classifier.in_features
        model.classifier = torch.nn.Linear(in_features=in_features, out_features=10)     
    if(model_name == 'inception_v3_MNIST'):
        model = inception_v3(weights = Inception_V3_Weights.DEFAULT)
        in_features = model.fc.in_features
        in_features2 = 768
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 10)
        model.AuxLogits = InceptionAux(in_channels = in_features2, num_classes = 10)         
    if(model_name == 'resnet50_IMGNET'):
        model = resnet50(weights = ResNet50_Weights.DEFAULT)
    if(model_name == 'resnet101_IMGNET'):
        model = resnet101(weights = ResNet101_Weights.DEFAULT)
    if(model_name == 'mobilenet_v2_IMGNET'):
        model = mobilenet_v2(weights = MobileNet_V2_Weights.DEFAULT)    
    if(model_name == 'alexnet_IMGNET'):
        model = alexnet(weights = AlexNet_Weights.DEFAULT)
    if(model_name == 'densenet121_IMGNET'):
        model = densenet121(weights = DenseNet121_Weights.DEFAULT)
    if(model_name == 'inception_v3_IMGNET'):
        model = inception_v3(weights = Inception_V3_Weights.DEFAULT)     
    if(path!=None and dataset_name!='imgnet'): 
        model.load_state_dict(torch.load(path, weights_only = True))
    return model.to(device)

def initData(path):
    if(dataset_name=='cifar100'):
        if(model_name=='inception_v3_CIFA100'): data_transforms = transforms_data_cifa100(width = 299, height = 299)
        else: data_transforms = transforms_data_cifa100()
        train_data = datasets.CIFAR100(
            root = path,
            train = True,
            download = True,
            transform = data_transforms['validation']
        )
        test_data = datasets.CIFAR100(
            root = path,
            train = False,
            download = True,
            transform = data_transforms['validation']
        )  
    elif(dataset_name=='cifar10'):
        if(model_name=='inception_v3_CIFA10'): data_transforms = transforms_data_cifa10(width = 299, height = 299)
        else: data_transforms = transforms_data_cifa10()
        train_data = datasets.CIFAR10(
            root = path,
            train = True,
            download = True,
            transform = data_transforms['validation']
        )
        test_data = datasets.CIFAR10(
            root = path,
            train = False,
            download = True,
            transform = data_transforms['validation']
        )
    elif(dataset_name=='mnist'):
        print(path)
        if(model_name=='inception_v3_MNIST'): data_transforms = transforms_data_mnist(width = 299, height = 299)
        else: data_transforms = transforms_data_mnist()
        train_data = datasets.MNIST(
            root = path,
            train = True,
            download = False,
            transform = data_transforms['validation']
        )
        test_data = datasets.MNIST(
            root = path,
            train = False,
            download = False,
            transform = data_transforms['validation']
        )
    elif(dataset_name=='imgnet'):
        if(model_name=='inception_v3_MNIST'): data_transforms = transforms_data_imgnet(width = 299, height = 299)
        else: data_transforms = transforms_data_imgnet()
        train_data = datasets.ImageNet(
            root = path,
            split = "val",
            transform = data_transforms['validation']
        )
        test_data = datasets.ImageNet(root = path, split = "val", transforms = data_transforms['validation'])
    train_loader = DataLoader(train_data, batch_size = bs, shuffle = False)
    test_loader = DataLoader(test_data, batch_size = bs, shuffle = False)
    return train_loader, test_loader
    
def initData2(path1, path2):
    train_loader = torch.load(path1)
    test_loader = torch.load(path2)
    return train_loader, test_loader

def pgd(model, X, y, criterion, lr, epsilon, epochs):
    delta = torch.zeros_like(X, requires_grad = True, device = device)
    for i in range(epochs):
        nX = X + delta
        loss = criterion(model(nX), y)
        loss.backward()
        delta.data = (delta + lr * delta.grad.data).clamp(-epsilon,epsilon)
        delta.grad.zero_()
    nX = X + delta.detach()
    return nX
    
def fgsm(model, imgs, labels, criterion, epsilon):
    delta = torch.zeros_like(imgs, requires_grad = True, device = device)
    outs = model(imgs+delta)
    loss=criterion(outs,labels)
    loss.backward()
    nimgs = imgs + epsilon * delta.grad.detach().sign()
    nimgs = torch.clamp(nimgs, 0, 1)
    return nimgs

def cw_attack_l2(model, imgs, labels, lr, step, c):
    atk = torchattacks.CW(model,c=c, steps=step, lr=lr)
    adv_imgs = atk(imgs,labels)
    return adv_imgs

def bim(model, imgs, labels, epsilon, alpha = 1/255):
    atk = torchattacks.BIM(model, eps = epsilon, alpha = alpha, steps = 0)
    adv_imgs = atk(imgs, labels)
    return adv_imgs

pgd_criterion = torch.nn.CrossEntropyLoss()
pgd_lr = 2/255
if(model_name[:9]=="inception"): pgd_lr = 2/299
pgd_epochs = 7

fgsm_criterion = torch.nn.CrossEntropyLoss()

bim_criterion = torch.nn.CrossEntropyLoss()
bim_alpha = 2/225
if(model_name[:9]="inception"): bim_alpha = 2/299

#cw_lr = 5e-4
cw_lr = 0.01
#cw_steps = 100
cw_steps = 4
#cw_c = 0.5
cw_c = 1

def adver_method(model, images, labels, epsilon, method):
    if(method == 'BIM'): 
        return bim(model, images, labels, epsilon, alpha = bim_alpha)
    if(method == 'FGSM'): 
        return fgsm(model, images, labels, fgsm_criterion, epsilon)
    if(method == 'PGD'): 
        return pgd(model, images, labels, pgd_criterion, pgd_lr, epsilon, pgd_epochs)
    if(method == 'CW_L2'):
        return cw_attack_l2(model, images, labels, cw_lr, cw_steps, cw_c)

def get_label_names(path):
    with open(path,"r") as file:
        data = json.load(file)
    return data

names = get_label_names(label_path)

def denormalize(img):
    new_img = img * torch.tensor(std).view(3, 1, 1) + torch.tensor(mean).view(3, 1, 1)
    return new_img

def Get_PIL_Image(img):
    img = denormalize(img)
    img = transforms.Resize((400, 400))(img)
    img = transforms.ToPILImage(mode='RGB')(img)
    return img

side = Side(style = 'thick')

def write_data(idx_row, sheet, out, gt_label):
    #print(torch.sum(out).item())
    pred = torch.argsort(out, dim = 0, descending=True)[:5]
    sheet.cell(row = idx_row, column = 1).value = names[int(gt_label)]
    sheet.cell(row = idx_row, column = 1).alignment = Alignment(horizontal='center')
    sheet.cell(row = idx_row, column = 1).border = Border(right = side)
    
    for i in range(0, 5):
        sheet.cell(row = idx_row + i, column = 2).value = names[pred[i]]
        sheet.cell(row = idx_row + i, column = 2).alignment = Alignment(horizontal='center')
        sheet.cell(row = idx_row + i, column = 2).border = Border(right = side, left = side)
        sheet.cell(row = idx_row + i, column = 3).value = out[pred[i]].item()
        sheet.cell(row = idx_row + i, column = 3).alignment = Alignment(horizontal='center')
        sheet.cell(row = idx_row + i, column = 3).border = Border(right = side, left = side)
    sheet.cell(row = idx_row + 4, column = 2).border = Border(right = side, left = side, bottom = side)
    sheet.cell(row = idx_row + 4, column = 3).border = Border(right = side, left = side, bottom = side)
    sheet.cell(row = idx_row + 4, column = 1).border = Border(right = side, left = side, bottom = side)
    
def Export_Result(idx, out1, out2, gt_label, book):
    sheet1 = book["original datasets"]
    write_data(idx, sheet1, out1, gt_label)
    sheet2 = book["adversarial datasets"]
    write_data(idx, sheet2, out2, gt_label)

def initWorkbook():
    book = yxl.Workbook()
    del book["Sheet"]
    for x in ["original datasets", "adversarial datasets"]:
        sheet = book.create_sheet(x)
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.merge_cells(start_row = 1, start_column = 2, end_row = 1, end_column = 3)
        sheet.merge_cells(start_row = 1, start_column = 1, end_row = 2, end_column = 1)
        sheet.cell(row = 1, column = 1).value = "Image True Class"
        sheet.cell(row = 1, column = 1).alignment = Alignment(horizontal='center', vertical = 'center')
        sheet.cell(row = 1, column = 2).value = "Top 5 Predicted Classes and Confidence"
        sheet.cell(row = 1, column = 2).alignment = Alignment(horizontal='center')
        sheet.cell(row = 2, column = 2).value = "Class"
        sheet.cell(row = 2, column = 2).alignment = Alignment(horizontal='center')
        sheet.cell(row = 2, column = 3).value = "Confidence"
        sheet.cell(row = 2, column = 3).alignment = Alignment(horizontal='center')
        sheet['A2'].border = Border(right = side, bottom = side)
        sheet['A1'].border = Border(right = side)
        sheet['B1'].border = Border(bottom = side,right = side)
        sheet['B2'].border = Border(bottom = side, right = side)
        sheet['C2'].border = Border(bottom = side, right = side)
        sheet['C1'].border = Border(bottom = side,right = side)
    return book

def initWorkbook2():
    book = yxl.Workbook()
    sheet = book["Sheet"]
    for i in range(0,5):
        sheet.column_dimensions[chr(ord('A')+i)].width = 30
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('B1:C1')
    sheet.merge_cells('D1:E1')

    sheet['A1'].value = 'Epsilon'
    sheet['B1'].value = 'Clean images'
    sheet['D1'].value = 'Adv. images'
    sheet['B2'].value = 'top-1'
    sheet['C2'].value = 'top-5'
    sheet['D2'].value = 'top-1'
    sheet['E2'].value = 'top-5'
    for dak in ['A1', 'B1', 'D1', 'B2', 'C2', 'D2', 'E2']:
        sheet[dak].alignment = Alignment(horizontal='center', vertical = 'center')
    return book

def get(model, loader, epsilon, method_name, book2, criterion):
    model.eval()
    print("{}: {}".format(epsilon, method_name))
    book = initWorkbook()
    excel_path = excel_dir+method_name+"/result_epsilon_{}.xlsx".format(epsilon)
    idx2 = 0 
    idx_excel_row = 3
    taken = np.zeros(1000)
    top_1_correct ={"ori": 0,"adv": 0}
    top_5_correct = {"ori": 0,"adv": 0}
    num_imgs = 0

    for batch_idx, (ori_imgs,labels) in enumerate(loader):
        if(batch_idx==5): break
        print("Batch: #{}".format(batch_idx))
        ori_imgs = ori_imgs.to(device)
        labels = labels.to(device)
        adv_imgs = adver_method(model, ori_imgs, labels, epsilon, method_name)
        out1s = F.softmax(model(ori_imgs), dim = 1)
        out2s = F.softmax(model(adv_imgs), dim = 1)
        pred={
            "ori":torch.argsort(out1s, dim = 1, descending = True),
            "adv":torch.argsort(out2s, dim = 1, descending = True)
        }
        for state in ["ori","adv"]:
            for idx2, dak in enumerate(pred[state]):
                top_1_correct[state] += (dak[0]==labels[idx2])
                top_5_correct[state] += (labels[idx2] in dak)
        num_imgs += ori_imgs.shape[0]

        for idx3 in range(0,bs):
            gt_label = labels[idx3].item()

            inp = ori_imgs[idx3][None,:].to(device)
            out1 = out1s[idx3]

            inp2 = adv_imgs[idx3][None,:].to(device)
            out2 = out2s[idx3]
            
            if(pred["ori"][idx3][0] != gt_label or taken[int(gt_label)]==1 or idx2==args.number_of_imgs):
                continue
            taken[int(gt_label)] = 1

            img_dst_dir_cur = os.path.join(img_dst_dir,method_name,"epsilon_{}".format(epsilon))
            make_directory([img_dst_dir_cur])
            path_adv_img = os.path.join(img_dst_dir_cur,"adv_img_{}.jpg".format(names[int(gt_label)]))
            path_ori_img = os.path.join(img_dst_dir_cur,"ori_img_{}.jpg".format(names[int(gt_label)]))
            path_noise_img = os.path.join(img_dst_dir_cur,"noise_img_{}.jpg".format(names[int(gt_label)]))    

            Export_Result(idx_excel_row, out1, out2, gt_label, book)

            inp = ori_imgs[idx3].to("cpu")
            img = Get_PIL_Image(inp)
            img.save(path_adv_img)

            inp2 = adv_imgs[idx3].to("cpu")
            img2 = Get_PIL_Image(inp2)
            img2.save(path_ori_img)

            inp3 = inp2-inp
            img3 = Get_PIL_Image(inp3)
            img3.save(path_noise_img)

            idx2 += 1
            idx_excel_row += 5
    book.save(excel_path)

    sheet = book2["Sheet"]
    sheet["A"+str(epsilon_idx+3)].value = epsilon
    sheet["A"+str(epsilon_idx+3)].alignment = Alignment(horizontal='center', vertical = 'center')
    for idx, state in enumerate(["ori", "adv"]):
        top_1 = top_1_correct[state] / num_imgs * 100
        top_5 = top_5_correct[state] / num_imgs * 100
        sheet[chr(ord('B')+idx*2)+str(epsilon_idx+3)].value = str(top_1.to("cpu").numpy())
        sheet[chr(ord('B')+idx*2)+str(epsilon_idx+3)].alignment = Alignment(horizontal='center', vertical = 'center')
        sheet[chr(ord('C')+idx*2)+str(epsilon_idx+3)].value = str(top_5)
        sheet[chr(ord('C')+idx*2)+str(epsilon_idx+3)].alignment = Alignment(horizontal='center', vertical = 'center')

    print("{} method has Correct / Total : {} / {}".format(method_name, top_1_correct["adv"], num_imgs))

if __name__=='__main__':
    nvmlInit()
    h = nvmlDeviceGetHandleByIndex(0)
    try:
        gpu_busy = True
        while(gpu_busy):
            info = nvmlDeviceGetMemoryInfo(h)
            print(f'used     : {int(info.used/(1.05*1000000))}',
            end="\r", flush=True)
            gpu_busy = int(info.used/(1.05*1000000)) > 6000
        system("cls")
                
        nnModel = initModel(model_path)
        criterion = nn.CrossEntropyLoss()
        
        train_loader, test_loader = initData(data_path)

        for method_name in adver_methods:
            make_directory([excel_dir+method_name+"/"])
            excel_path_2 = excel_dir+method_name+"/top_1_top_5.xlsx"
            book2 =  initWorkbook2()
            for epsilon_idx, epsilon in enumerate([0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09, 0.1]):
                if(method_name == 'CW_L2' and epsilon_idx>0): continue
                dst_test_dir = dst_dir + method_name+"/test_dataset/"
                dst_train_dir = dst_dir + method_name + "/train_dataset/"
                make_directory([dst_test_dir])
                ad_test_path = dst_test_dir+"adversarial_test_{}.pt".format(epsilon_idx+1)
                ad_train_path = dst_train_dir+"adversarial_train_{}.pt".format(epsilon_idx+1)
                ori_test_path = dst_test_dir+"original_test_{}.pt".format(epsilon_idx+1)
                ori_train_path = dst_train_dir+"original_train_{}.pt".format(epsilon_idx+1)

                #ad_train, ori_train = adversarial_attacks("Train",nnModel, train_loader, criterion, method_name, epsilon = epsilon)
                #torch.save(ad_train, ad_train_path)          
                #torch.save(ori_train,ori_train_path)
                get(nnModel, test_loader, epsilon, method_name, book2, criterion)
            
            book2.save(excel_path_2)


    except KeyboardInterrupt:
        print("Press Ctrl-C to terminate while statement")
    
