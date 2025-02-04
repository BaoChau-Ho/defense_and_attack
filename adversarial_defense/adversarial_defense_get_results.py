import torch
import argparse
import sys
from os import system
import time
import copy
import numpy as np
from torch import optim
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader, TensorDataset
from torchvision import datasets
from torchvision.models.inception import InceptionAux
from torchvision.models import mobilenet_v2, resnet50, alexnet, resnet101, densenet121, inception_v3, DenseNet121_Weights, ResNet101_Weights, ResNet50_Weights, MobileNet_V2_Weights, AlexNet_Weights, Inception_V3_Weights
from torchvision import transforms
from pytorchtools import EarlyStopping
from utils import Logger
from pynvml import *

import json
import os
import openpyxl as yxl
import numpy as np
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

import subprocess
subprocess.run(["pip3","install","torchattacks"])
import torchattacks

def make_directory(dirs):
    for dir in dirs:
        if(not os.path.exists(dir)):
            os.makedirs(dir)

device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
device1 = torch.device("cuda:0")

parser = argparse.ArgumentParser()
parser.add_argument("--index_model",type=int)
parser.add_argument("--index_training",type=int)
args=parser.parse_args()

father_directory = "/media/administrator/Data1/BC/Test_Python/"
#father_directory = "/home/kt/bc/adversarial_resnet50_cifar100/"
#father_directory = "/mnt/data/dtg/giangdt/BC/adversarial_robustness/"

models = ['resnet50_CIFA100', 'resnet101_CIFA100', 'mobilenet_v2_CIFA100', 'alexnet_CIFA100', 'densenet121_CIFA100','inception_v3_CIFA100',
          'resnet50_CIFA10', 'resnet101_CIFA10', 'mobilenet_v2_CIFA10', 'alexnet_CIFA10', 'densenet121_CIFA10', 'inception_v3_CIFA10',
          'resnet50_MNIST', 'resnet101_MNIST', 'mobilenet_v2_MNIST', 'alexnet_MNIST', 'densenet121_MNIST', 'inception_v3_MNIST',
          'resnet50_IMGNET', 'resnet101_IMGNET','mobilenet_v2_IMGNET','alexnet_IMGNET','densenet121_IMGNET','inception_v3_IMGNET']
adver_methods = ['FGSM','BIM','PGD','CW_L2']
all_datasets = ['cifar100','cifar10','mnist','imgnet']
training_methods = ["ATWR", "ATGR", "EATR"]

bs={
    "cifar100": 32,
    "cifar10": 32,
    "mnist": 32,
    "imgnet": 32
}

#adver_method_name = adver_methods[2]
# 0-5: cifar100
# 6-11: cifar10
# 12-17: mnist
# 18-23: imgnet
# /home/kt/bc/

def get_dataset_index(model_idx):
    if(model_idx<6): dak=0
    elif(model_idx<12): dak=1
    elif(model_idx<18): dak=2
    elif(model_idx<24): dak=3
    return dak

model_name = models[args.index_model] 
dataset_name = all_datasets[get_dataset_index(args.index_model)]
training_method_name = training_methods[args.index_training]

model_path = father_directory + "models/{}/{}/weights_{}_best.h5".format(training_method_name, model_name, model_name)
data_path = father_directory + "datasets_{}/".format(dataset_name)
excel_dir = father_directory + "trained_model_results/{}/".format(training_method_name)
print(model_name)
print(dataset_name)

make_directory([excel_dir])

print(excel_dir)

mean = (0.4914, 0.4822, 0.4465)
std = (0.247, 0.243, 0.261)
    
def transforms_data_cifa100(width=224, height=224):
    transforms_dict = {
        'validation':
        transforms.Compose([
            transforms.Resize((width, height)),
            transforms.ToTensor(),
            transforms.Normalize(mean, std)
        ]),
    }
    return transforms_dict
    
def transforms_data_mnist(width=224, height=224):
    transforms_dict = {
        'validation':
        transforms.Compose([
            transforms.Grayscale(3),
            transforms.Resize((width, height)),
            transforms.ToTensor(),
            transforms.Normalize(mean, std)
        ]),
    }
    return transforms_dict

def transforms_data_cifa10(width=224, height=224):
    transforms_dict = {
        'validation':
        transforms.Compose([
            transforms.Resize((width, height)),
            transforms.ToTensor(),
            transforms.Normalize(mean, std)
        ])
    }
    return transforms_dict

def transforms_data_imgnet():
    if(model_name=='inception_v3_MNIST'):
        data_transforms = {
            "validation": transforms.Compose([
                transforms.Resize((299, 299)),
                transforms.ToTensor(), # ToTensor : [0, 255] -> [0, 1]
                transforms.Normalize(mean, std)])
        }        
    else:
        # models except for inception
        transforms_dict = {
            'validation':
            transforms.Compose([
                transforms.Resize(256),
                transforms.CenterCrop(224),
                transforms.ToTensor(),
                transforms.Normalize(mean, std)
            ])
        }
    return transforms_dict

def initModel(path = None):
    if(model_name == 'mobilenet_v2_CIFA100'):
        model = mobilenet_v2(weights = MobileNet_V2_Weights.DEFAULT)
        in_features = model.classifier[1].in_features
        model.classifier[1] = torch.nn.Linear(in_features=in_features, out_features=100)
    if(model_name == 'resnet50_CIFA100'):
        model = resnet50(weights = ResNet50_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features=in_features, out_features = 100)
    if(model_name == 'alexnet_CIFA100'):
        model = alexnet(weights = AlexNet_Weights.DEFAULT)
        in_features = model.classifier[6].in_features
        model.classifier[6] = torch.nn.Linear(in_features = in_features, out_features = 100)
    if(model_name == 'resnet101_CIFA100'):
        model = resnet101(weights = ResNet101_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 100)
    if(model_name == 'densenet121_CIFA100'):
        model = densenet121(weights = DenseNet121_Weights.DEFAULT)
        in_features = model.classifier.in_features
        model.classifier = torch.nn.Linear(in_features=in_features, out_features=100)
    if(model_name == 'inception_v3_CIFA100'):
        model = inception_v3(weights = Inception_V3_Weights.DEFAULT)
        in_features = model.fc.in_features
        in_features2 = 768
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 100)
        model.AuxLogits = InceptionAux(in_channels = in_features2, num_classes = 100)
    if(model_name == 'resnet50_CIFA10'):
        model = resnet50(weights = ResNet50_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features=in_features, out_features = 10)
    if(model_name == 'resnet101_CIFA10'):
        model = resnet101(weights = ResNet101_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 10)
    if(model_name == 'mobilenet_v2_CIFA10'):
        model = mobilenet_v2(weights = MobileNet_V2_Weights.DEFAULT)
        in_features = model.classifier[1].in_features
        model.classifier[1] = torch.nn.Linear(in_features=in_features, out_features=10)       
    if(model_name == 'alexnet_CIFA10'):
        model = alexnet(weights = AlexNet_Weights.DEFAULT)
        in_features = model.classifier[6].in_features
        model.classifier[6] = torch.nn.Linear(in_features = in_features, out_features = 10) 
    if(model_name == 'densenet121_CIFA10'):
        model = densenet121(weights = DenseNet121_Weights.DEFAULT)
        in_features = model.classifier.in_features
        model.classifier = torch.nn.Linear(in_features=in_features, out_features=10)     
    if(model_name == 'inception_v3_CIFA10'):
        model = inception_v3(weights = Inception_V3_Weights.DEFAULT)
        in_features = model.fc.in_features
        in_features2 = 768
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 10)
        model.AuxLogits = InceptionAux(in_channels = in_features2, num_classes = 10)  
    if(model_name == 'resnet50_MNIST'):
        model = resnet50(weights = ResNet50_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features=in_features, out_features = 10)
    if(model_name == 'resnet101_MNIST'):
        model = resnet101(weights = ResNet101_Weights.DEFAULT)
        in_features = model.fc.in_features
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 10)
    if(model_name == 'mobilenet_v2_MNIST'):
        model = mobilenet_v2(weights = MobileNet_V2_Weights.DEFAULT)
        in_features = model.classifier[1].in_features
        model.classifier[1] = torch.nn.Linear(in_features=in_features, out_features=10)       
    if(model_name == 'alexnet_MNIST'):
        model = alexnet(weights = AlexNet_Weights.DEFAULT)
        in_features = model.classifier[6].in_features
        model.classifier[6] = torch.nn.Linear(in_features = in_features, out_features = 10) 
    if(model_name == 'densenet121_MNIST'):
        model = densenet121(weights = DenseNet121_Weights.DEFAULT)
        in_features = model.classifier.in_features
        model.classifier = torch.nn.Linear(in_features=in_features, out_features=10)     
    if(model_name == 'inception_v3_MNIST'):
        model = inception_v3(weights = Inception_V3_Weights.DEFAULT)
        in_features = model.fc.in_features
        in_features2 = 768
        model.fc = torch.nn.Linear(in_features = in_features, out_features = 10)
        model.AuxLogits = InceptionAux(in_channels = in_features2, num_classes = 10)         
    if(model_name == 'resnet50_IMGNET'):
        model = resnet50(weights = ResNet50_Weights.IMAGENET1K_V2)
    if(model_name == 'resnet101_IMGNET'):
        model = resnet101(weights = ResNet101_Weights.DEFAULT)
    if(model_name == 'mobilenet_v2_IMGNET'):
        model = mobilenet_v2(weights = MobileNet_V2_Weights.DEFAULT)    
    if(model_name == 'alexnet_IMGNET'):
        model = alexnet(weights = AlexNet_Weights.DEFAULT)
    if(model_name == 'densenet121_IMGNET'):
        model = densenet121(weights = DenseNet121_Weights.DEFAULT)
    if(model_name == 'inception_v3_IMGNET'):
        model = inception_v3(weights = Inception_V3_Weights.DEFAULT)     
    if(path!=None and dataset_name!='imgnet'): 
        model.load_state_dict(torch.load(path, weights_only = True))
    return model.to(device)

def initData(path):
    if(dataset_name=='cifar100'):
        if(model_name[:12]=="inception_v3"): data_transforms = transforms_data_cifa100(width = 299, height = 299)
        else: data_transforms = transforms_data_cifa100()
        test_data = datasets.CIFAR100(
            root = path,
            train = False,
            download = True,
            transform = data_transforms['validation']
        )  
    elif(dataset_name=='cifar10'):
        if(model_name[:12]=="inception_v3"): data_transforms = transforms_data_cifa10(width = 299, height = 299)
        else: data_transforms = transforms_data_cifa10()
        test_data = datasets.CIFAR10(
            root = path,
            train = False,
            download = True,
            transform = data_transforms['validation']
        )
    elif(dataset_name=='mnist'):
        print(path)
        if(model_name[:12]=="inception_v3"): data_transforms = transforms_data_mnist(width = 299, height = 299)
        else: data_transforms = transforms_data_mnist()
        test_data = datasets.MNIST(
            root = path,
            train = False,
            download = True,
            transform = data_transforms['validation']
        )
    elif(dataset_name=='imgnet'):
        if(model_name[:12]=="inception_v3"): data_transforms = transforms_data_imgnet(width = 299, height = 299)
        else: data_transforms = transforms_data_imgnet()
        test_data = datasets.ImageNet(root = path, split = "val", transform = data_transforms['validation'])
    test_loader = DataLoader(test_data, batch_size = bs[dataset_name], shuffle = False)
    return test_loader

def pgd(model, X, y, criterion, lr, epsilon, epochs):
    delta = torch.zeros_like(X, requires_grad = True, device = device)
    for i in range(epochs):
        nX = X + delta
        loss = criterion(model(nX), y)
        loss.backward()
        delta.data = (delta + lr * delta.grad.data).clamp(-epsilon,epsilon)
        delta.grad.zero_()
    nX = X + delta.detach()
    return nX
    
def fgsm(model, imgs, labels, criterion, epsilon):
    delta = torch.zeros_like(imgs, requires_grad = True, device = device)
    outs = model(imgs+delta)
    loss=criterion(outs,labels)
    loss.backward()
    nimgs = imgs + epsilon * delta.grad.detach().sign()
    nimgs = torch.clamp(nimgs, 0, 1)
    return nimgs

def cw_attack_l2(model, imgs, labels, lr, bs_step):
    atk = torchattacks.CW(model, steps=bs_step, lr=lr)
    adv_imgs = atk(imgs,labels)
    return adv_imgs

def bim(model, imgs, labels, epsilon, alpha = 1/255):
    atk = torchattacks.BIM(model, eps = epsilon, alpha = alpha, steps = 0)
    adv_imgs = atk(imgs, labels)
    return adv_imgs

pgd_criterion = torch.nn.CrossEntropyLoss()
pgd_lr = 2/255
pgd_epsilon = 0.1
pgd_epochs = 7

fgsm_criterion = torch.nn.CrossEntropyLoss()
fgsm_epsilon = 0.1

bim_criterion = torch.nn.CrossEntropyLoss()
bim_epsilon = 0.1
bim_alpha = 1/225

cw_lr = 0.01
cw_bs_step = 4

def adver_method(model, images, labels, method):
    if(method == 'BIM'): 
        return bim(model, images, labels, bim_epsilon, alpha = bim_alpha)
    if(method == 'FGSM'): 
        return fgsm(model, images, labels, fgsm_criterion, fgsm_epsilon)
    if(method == 'PGD'): 
        return pgd(model, images, labels, pgd_criterion, pgd_lr, pgd_epsilon, pgd_epochs)
    if(method == 'CW_L2'):
        return cw_attack_l2(model, images, labels, cw_lr, cw_bs_step)
    
def initWorkbook():
    book = yxl.Workbook()
    sheet = book["Sheet"]
    for i in range(0,5):
        sheet.column_dimensions[chr(ord('A')+i)].width = 30
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('B1:C1')
    sheet.merge_cells('D1:E1')

    sheet['A1'].value = 'Adv method'
    sheet['B1'].value = 'Clean images'
    sheet['D1'].value = 'Adv. images'
    sheet['B2'].value = 'top-1'
    sheet['C2'].value = 'top-5'
    sheet['D2'].value = 'top-1'
    sheet['E2'].value = 'top-5'
    for dak in ['A1', 'B1', 'D1', 'B2', 'C2', 'D2', 'E2']:
        sheet[dak].alignment = Alignment(horizontal='center', vertical = 'center')
    return book

def get(model, loader, method_idx, method_name, book):
    model.eval()
    print("{}".format(method_name))
    top_1_correct ={"ori": 0,"adv": 0}
    top_5_correct = {"ori": 0,"adv": 0}
    num_imgs = 0

    for batch_idx, (ori_imgs,labels) in enumerate(loader):
        if(batch_idx==10): break
        ori_imgs = ori_imgs.to(device)
        labels = labels.to(device)
        adv_imgs = adver_method(model, ori_imgs, labels, method_name)
        out1s = model(ori_imgs)
        out2s = model(adv_imgs)
        pred={
            "ori":torch.topk(out1s, 5, 1, True, True)[1],
            "adv":torch.topk(out2s, 5, 1, True, True)[1]
        }
        for state in ["ori","adv"]:
            for idx2, dak in enumerate(pred[state]):
                top_1_correct[state] += (dak[0]==labels[idx2])
                top_5_correct[state] += (labels[idx2] in dak)
        num_imgs += ori_imgs.shape[0]

    sheet = book["Sheet"]
    sheet["A"+str(method_idx+3)].value = method_name
    sheet["A"+str(method_idx+3)].alignment = Alignment(horizontal='center', vertical = 'center')
    for idx, state in enumerate(["ori", "adv"]):
        top_1 = top_1_correct[state] / num_imgs * 100
        top_5 = top_5_correct[state] / num_imgs * 100
        sheet[chr(ord('B')+idx*2)+str(method_idx+3)].value = "{:.2f}".format(top_1)
        sheet[chr(ord('B')+idx*2)+str(method_idx+3)].alignment = Alignment(horizontal='center', vertical = 'center')
        sheet[chr(ord('C')+idx*2)+str(method_idx+3)].value = "{:.2f}".format(top_5)
        sheet[chr(ord('C')+idx*2)+str(method_idx+3)].alignment = Alignment(horizontal='center', vertical = 'center')

    print("{} method has Correct / Total : {} / {}".format(method_name, top_1_correct["adv"], num_imgs))

if __name__=='__main__':
    nvmlInit()
    h = nvmlDeviceGetHandleByIndex(0)
    try:
        gpu_busy = True
        while(gpu_busy):
            info = nvmlDeviceGetMemoryInfo(h)
            print(f'used     : {int(info.used/(1.05*1000000))}',
            end="\r", flush=True)
            gpu_busy = int(info.used/(1.05*1000000)) > 7000
        system("cls")
                
        nnModel = initModel(model_path)
        criterion = nn.CrossEntropyLoss()
        
        test_loader = initData(data_path)
        excel_path = excel_dir+"{}_{}_top_1_top_5.xlsx".format(model_name,training_method_name)
        book =  initWorkbook()
        for method_idx, method_name in enumerate(adver_methods):
            get(nnModel, test_loader, method_idx, method_name, book)
        book.save(excel_path)

    except KeyboardInterrupt:
        print("Press Ctrl-C to terminate while statement")
    
